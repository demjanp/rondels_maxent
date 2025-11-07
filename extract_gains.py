"""
Utility script to summarize MaxEnt training gains across scenario folders.

Usage:
    python extract_gains.py <input-directory>
"""
from __future__ import annotations

import argparse
import csv
import sys
from pathlib import Path
from typing import Dict, Iterable, List, Tuple, Union

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

WITHOUT_PREFIX_RAW = "Training gain without "
WITH_ONLY_PREFIX_RAW = "Training gain with only "
WITHOUT_PREFIX = WITHOUT_PREFIX_RAW.lower()
WITH_ONLY_PREFIX = WITH_ONLY_PREFIX_RAW.lower()
OUTPUT_HEADERS = ["Variable", "Training gain without", "Training gain with only"]
ValueType = Union[str, float]


class ExtractionError(Exception):
    """Raised when the expected CSV structure is not present."""


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Extract training gain metrics from MaxEnt result folders and create XLSX tables."
    )
    parser.add_argument(
        "input_path",
        type=Path,
        help=(
            "Path containing scenario subdirectories. Each subdirectory must "
            "include maxent/maxentResults.csv."
        ),
    )
    return parser.parse_args()


def read_training_gains(csv_path: Path) -> Tuple[Dict[str, str], Dict[str, str]]:
    """Read a MaxEnt results CSV and split training gain metrics into buckets."""
    try:
        with csv_path.open(newline="", encoding="utf-8") as handle:
            reader = csv.reader(handle)
            headers = next(reader)
            values = next(reader)
    except FileNotFoundError as exc:
        raise ExtractionError(f"Missing file: {csv_path}") from exc
    except StopIteration as exc:
        raise ExtractionError(f"Incomplete CSV: {csv_path}") from exc

    # Pad values to match headers if necessary (some CSV tools drop trailing empties).
    if len(values) < len(headers):
        values += [""] * (len(headers) - len(values))

    without: Dict[str, str] = {}
    with_only: Dict[str, str] = {}

    for header, value in zip(headers, values):
        if not header:
            continue
        stripped_header = header.strip()
        normalized_header = stripped_header.lower()
        normalized_value = value.strip()

        if normalized_header.startswith(WITHOUT_PREFIX):
            variable = stripped_header[len(WITHOUT_PREFIX_RAW) :].strip()
            without[variable] = normalized_value
        elif normalized_header.startswith(WITH_ONLY_PREFIX):
            variable = stripped_header[len(WITH_ONLY_PREFIX_RAW) :].strip()
            with_only[variable] = normalized_value

    return without, with_only


def build_rows(
    without: Dict[str, str], with_only: Dict[str, str]
) -> List[Dict[str, str]]:
    """Combine both dictionaries into table rows keyed by variable name."""
    variables = sorted(set(without) | set(with_only))
    rows: List[Dict[str, str]] = []
    for variable in variables:
        rows.append(
            {
                "Variable": variable,
                "Training gain without": without.get(variable, ""),
                "Training gain with only": with_only.get(variable, ""),
            }
        )
    return rows


def _coerce_number(value: str) -> ValueType:
    """Convert numeric-looking strings to floats so Excel treats them as numbers."""
    if value == "":
        return ""
    try:
        return float(value)
    except ValueError:
        return value


def write_xlsx(rows: Iterable[Dict[str, str]], destination: Path) -> None:
    """Persist rows into a real XLSX file using openpyxl."""
    destination.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Training Gains"

    worksheet.append(OUTPUT_HEADERS)
    for row in rows:
        worksheet.append(
            [
                row.get("Variable", ""),
                _coerce_number(row.get("Training gain without", "")),
                _coerce_number(row.get("Training gain with only", "")),
            ]
        )

    # Autosize columns for readability.
    for idx, heading in enumerate(OUTPUT_HEADERS, start=1):
        column_values = [heading] + [row.get(heading, "") for row in rows]
        max_length = max((len(str(value)) for value in column_values), default=len(heading))
        worksheet.column_dimensions[get_column_letter(idx)].width = min(max_length + 2, 50)

    workbook.save(destination)


def process_directory(input_root: Path) -> None:
    any_written = False
    for child in sorted(path for path in input_root.iterdir() if path.is_dir()):
        csv_path = child / "maxent" / "maxentResults.csv"
        if not csv_path.exists():
            print(f"[skip] {csv_path} not found", file=sys.stderr)
            continue

        try:
            without, with_only = read_training_gains(csv_path)
        except ExtractionError as exc:
            print(f"[error] {exc}", file=sys.stderr)
            continue

        rows = build_rows(without, with_only)
        if not rows:
            print(f"[warn] No training gain fields found in {csv_path}", file=sys.stderr)
            continue

        destination = input_root / f"{child.name}.xlsx"
        write_xlsx(rows, destination)
        any_written = True
        print(f"[ok] Wrote {destination}")

    if not any_written:
        raise SystemExit("No valid maxent results found.")


def main() -> None:
    args = parse_args()
    input_root: Path = args.input_path

    if not input_root.exists():
        raise SystemExit(f"Input path does not exist: {input_root}")
    if not input_root.is_dir():
        raise SystemExit(f"Input path must be a directory: {input_root}")

    process_directory(input_root)


if __name__ == "__main__":
    main()
